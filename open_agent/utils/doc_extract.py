"""Binary document text extraction for PDF, DOCX, and XLSX files.

Provides a unified interface to extract text from binary document formats.
Each extractor degrades gracefully when its library is not installed.

Dependencies (all optional):
- PyMuPDF (fitz) for PDF
- python-docx for DOCX
- openpyxl for XLSX
"""

from pathlib import Path


# Binary document extensions that need special handling
BINARY_DOC_EXTS = frozenset({".pdf", ".docx", ".doc", ".xlsx", ".xls"})

# Maximum content size to prevent memory issues
MAX_FILE_CONTENT = 2_000_000  # 2MB


def _truncate(text: str) -> str:
    """Truncate text if it exceeds the maximum size."""
    if len(text) > MAX_FILE_CONTENT:
        return (
            text[:MAX_FILE_CONTENT]
            + "\n\n[...文件过大，已截断至前 2MB。请使用 offset/limit 分段读取。]"
        )
    return text


def extract_pdf_text(file_path: Path) -> str:
    """Extract text from a PDF file using PyMuPDF.

    Args:
        file_path: Path to the PDF file.

    Returns:
        Extracted text content.

    Raises:
        ImportError: If PyMuPDF is not installed.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return "[PDF 提取需要 PyMuPDF 库。请运行: pip install PyMuPDF]"

    try:
        doc = fitz.open(str(file_path))
        pages = []
        for page_num, page in enumerate(doc, 1):
            text = page.get_text().strip()
            if text:
                pages.append(f"--- 第 {page_num} 页 ---\n{text}")
        doc.close()
        return _truncate("\n\n".join(pages)) if pages else "[PDF 未提取到文本内容]"
    except Exception as e:
        return f"[PDF 提取失败: {e}]"


def extract_docx_text(file_path: Path) -> str:
    """Extract text from a DOCX file using python-docx.

    Args:
        file_path: Path to the DOCX file.

    Returns:
        Extracted text content.

    Raises:
        ImportError: If python-docx is not installed.
    """
    try:
        import docx
    except ImportError:
        return "[DOCX 提取需要 python-docx 库。请运行: pip install python-docx]"

    try:
        doc = docx.Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return _truncate("\n".join(paragraphs)) if paragraphs else "[DOCX 未提取到文本内容]"
    except Exception as e:
        return f"[DOCX 提取失败: {e}]"


def extract_xlsx_text(file_path: Path) -> str:
    """Extract text from an XLSX/XLS file using openpyxl.

    Args:
        file_path: Path to the spreadsheet file.

    Returns:
        Extracted text content with sheet separators.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError:
        return "[XLSX 提取需要 openpyxl 库。请运行: pip install openpyxl]"

    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                sheets.append(f"--- {sheet_name} ---\n" + "\n".join(rows))
        wb.close()
        return _truncate("\n\n".join(sheets)) if sheets else "[XLSX 未提取到数据]"
    except Exception as e:
        return f"[XLSX 提取失败: {e}]"


def read_file_content(file_path: Path) -> str:
    """Read content from a file, dispatching to the appropriate extractor.

    For binary documents (PDF, DOCX, XLSX), uses specialized extractors.
    For text files, reads directly.

    Args:
        file_path: Path to the file.

    Returns:
        File content as text.
    """
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        return extract_pdf_text(file_path)
    elif ext == ".docx":
        return extract_docx_text(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx_text(file_path)
    elif ext == ".doc":
        # .doc (legacy format) is not supported without additional libraries
        return "[.doc 格式暂不支持。请转换为 .docx 格式。]"
    else:
        # Text file — read directly
        try:
            return _truncate(file_path.read_text(encoding="utf-8", errors="replace"))
        except Exception as e:
            return f"[文件读取失败: {e}]"


def is_binary_doc(file_path: Path) -> bool:
    """Check if a file is a binary document that needs special extraction."""
    return file_path.suffix.lower() in BINARY_DOC_EXTS
